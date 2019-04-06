from openpyxl import Workbook, utils, worksheet
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, named_styles, colors
from XMLReader import XMLReader

import time
from babel.dates import format_date
from datetime import date

class excel_forma:
    # создание объекта
    def __init__(self, path_xml, path_excel):
        self._wb = Workbook()
        self.sec_since_epoch = list(time.localtime())[:3][::-1]

        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)
        self.font = Font(bold=True, color="000000")
        self.al_center = Alignment(horizontal="center", vertical="center", wrapText = True)
        self.al_left = Alignment(horizontal="left", vertical="center", wrapText = True)

        self._TITLE_FRS_PAGE_ROW = 1 # 1 row
        self._TITLE_FRS_PAGE_COL = 2 # 'B' column

        self._TITLE_SCN_PAGE_ROW = self._TITLE_FRS_PAGE_ROW + 2 # 3 row
        self._TITLE_SCN_PAGE_COL = self._TITLE_FRS_PAGE_COL     # 'B' column

        self._TITLE_THR_PAGE_ROW = self._TITLE_SCN_PAGE_ROW + 1 # 4 row
        self._TITLE_THR_PAGE_COL = self._TITLE_FRS_PAGE_COL     # 'B' column

        self._DATE_ROW = self._TITLE_FRS_PAGE_ROW       # 1 row
        self._DATE_COL = self._TITLE_FRS_PAGE_COL + 5   # 'G' column

        self._START_ROW_TABLE = self._TITLE_FRS_PAGE_ROW + 6    # 7 row
        self._START_COL_TABLE = self._TITLE_FRS_PAGE_COL - 1    # 'A' column



        self.path_xml = path_xml
        self.path_excel = path_excel
    #формирование книги
    def book_forming(self):
        self._ws_page2 = self._wb.active
        self._ws_page2.title = "Page"

        data_price = self._data_page2_to_dict()

        self._merge_page2_cells()
        self._create_page(data_price)
        self.save_file()
    #чтение из файла
    def _reading_file(self, path):
        ls_price = XMLReader(path).get_list_of_dicts()
        if ls_price:
            return ls_price 
        else:
            raise Exception("File " + path + " is empty ")
    #форматирование ячеек
    def _merge_page2_cells(self):
        self._ws_page2.column_dimensions['A'].width = 5.43
        self._ws_page2.column_dimensions['B'].width = 35.57
        self._ws_page2.column_dimensions['C'].width = 12.29
        self._ws_page2.column_dimensions['D'].width = 12.29
        self._ws_page2.column_dimensions['E'].width = 22.71
        self._ws_page2.column_dimensions['G'].width = 23

        self._ws_page2.row_dimensions[1].height = 33
        self._ws_page2.row_dimensions[3].height = 33

        self._select_styles_page(self._ws_page2, "B1:E1", font = self.font, alignment = self.al_center)
        self._select_styles_page(self._ws_page2, "B3:E3", font = self.font, alignment = self.al_center)
        self._select_styles_page(self._ws_page2, "B4:E4", font = self.font, alignment = self.al_center)


        self._ws_page2["G1"].alignment = self.al_left        
        self._ws_page2["G1"].font = self.font
    #установка стилей
    def _select_styles_page(self, ws, cell_range, border=None, font=None, alignment=None):
        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = list(ws[cell_range])

        if font:
            first_cell.font = font

        if border:
            for cell in rows[0]:
                cell.border = border
            for cell in rows[-1]:
                cell.border = border

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = border
                r.border = border
	
    def _filling_cell(self, dt):

        self._ws_page2.cell(row = self._TITLE_FRS_PAGE_ROW, 
                            column = self._TITLE_FRS_PAGE_COL, 
                            value = "Отчет сгенерирован Функциональной подсистемой «Управление сельского хозяйства» акимата города Астаны")
        self._ws_page2.cell(row = self._DATE_ROW, 
                            column = self._DATE_COL, 
                            value  = "Дата выдачи: " + ".".join(str(el) for el in self.sec_since_epoch))
        self._ws_page2.cell(row = self._TITLE_SCN_PAGE_ROW, 
                            column = self._TITLE_SCN_PAGE_COL, 
                            value = "Еженедельный отчет о мониторинге цен на социально-значимые продовольственные товары")
        self._ws_page2.cell(row = self._TITLE_THR_PAGE_ROW, 
                            column = self._TITLE_THR_PAGE_COL, 
                            value = "отчетный период за " + format_date(date(int(dt[0]), int(dt[1]), int(dt[2])), format = 'long', locale = 'ru_RU'))
    # создание листа
    def _create_page(self, data_price):
        dt = self.date.split('T')
        dt = dt[0].split('-')
        self._filling_cell(dt)
        self._create_table(data_price, dt)

    # заполнение листа
    def _create_table(self, data_price, dt):

        dt_beg = list(time.gmtime(time.mktime((int(dt[0]), int(dt[1]), int(dt[2]), 0, 0, 0, 0, 0, 0)) - 6*24*60*60))[:3]
        dt_beg = '.'.join(str(el) for el in dt_beg[::-1])
        dt_end = '.'.join(str(el) for el in dt[::-1])

        self._ws_page2.cell(row = self._START_ROW_TABLE, 
                            column = self._START_COL_TABLE, 
                            value = "№")

        self._ws_page2.cell(row = self._START_ROW_TABLE, 
                            column = self._START_COL_TABLE + 1, 
                            value = "Наименование социально-значимых продовольственных товаров")

        self._ws_page2.cell(row = self._START_ROW_TABLE, 
                            column = self._START_COL_TABLE + 2, 
                            value = "Цена за " + dt_beg)

        self._ws_page2.cell(row = self._START_ROW_TABLE, 
                            column = self._START_COL_TABLE + 3, 
                            value = "Цена за " + dt_end)

        self._ws_page2.cell(row = self._START_ROW_TABLE, 
                            column = self._START_COL_TABLE + 4, 
                            value = "сравнение в %, " + dt_beg + " с " + dt_end)

        product_names = self._reading_file(self.path_xml[2])

        ls_product_names = []
        idx_name_product = 1

        for array in product_names:
            ls_product_names.append(array['attrib']['fullname'])
            self._ws_page2.cell(row = 8 + idx_name_product, column = 1, value = idx_name_product)
            self._ws_page2.cell(row = 8 + idx_name_product, column = 2, value = ls_product_names[-1])            
            idx_name_product += 1

        idx = 1

        for col_letter in utils.get_column_interval(1, 5):
            self._ws_page2[col_letter + '8'] = idx
            idx += 1
            for row_number in range(len(ls_product_names) + 2):
                self._ws_page2[col_letter + str(7 + row_number)].border = self.border
                
                self._ws_page2[col_letter + str(7 + row_number)].font = self.font
                if col_letter == 'B' and 9 <= 7 + row_number:
                    self._ws_page2[col_letter + str(7 + row_number)].alignment = self.al_left
                else:
                    self._ws_page2[col_letter + str(7 + row_number)].alignment = self.al_center
                
        self._filling_table(data_price, ls_product_names)

    # данные с xml преобразовать в (ключ, значение)
    def _data_page2_to_dict(self):
        ls_price_now_week = self._reading_file(self.path_xml[0])

        self.date = ls_price_now_week[0]['attrib']['Dend']
        
        ls_price_last_week = self._reading_file(self.path_xml[1])

        dict_prod_names = {}
        idx = 0
        for ls_now_week in ls_price_now_week:
            product_name_now_week = ls_now_week['attrib']['item2']
            for ls_last_week in ls_price_last_week:
                if ls_last_week['attrib']['item2'] == product_name_now_week:
                    price_last = int(float(ls_last_week['attrib']['Value1']))
                    price_now = int(float(ls_now_week['attrib']['Value1']))
                    procent = round((price_now / float(price_last)) * 100 - 100, 2)
                    dict_prod_names[product_name_now_week] = [price_last, price_now, procent]
                    break

            if dict_prod_names.get(product_name_now_week) is None:
                dict_prod_names[product_name_now_week] = ['', '', '']
        return dict_prod_names

    # заполнение таблицы
    def _filling_table(self, dict_price, ls_product_names):
            
        for key in dict_price.keys():
            try:
                i_row = ls_product_names.index(key) + 9
            except ValueError:
                pass

            self._ws_page2.cell(row=i_row, column = 3, value = dict_price[key][0])
            self._ws_page2.cell(row=i_row, column = 4, value = dict_price[key][1])
            if dict_price[key][2] != '':
                if dict_price[key][2] <= 0:
                    self._ws_page2.cell(row=i_row, column = 5, value = dict_price[key][2]).fill = PatternFill(patternType = 'solid', fgColor = colors.GREEN)
                else:
                    self._ws_page2.cell(row=i_row, column = 5, value = dict_price[key][2]).fill = PatternFill(patternType = 'solid', fgColor = colors.RED)
    # сохранение в файл
    def save_file(self): 
        self._wb.save(self.path_excel)
        print("true")