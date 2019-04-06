from openpyxl import Workbook, utils, worksheet
from openpyxl.styles import Fill, Font, Border, Side, Alignment, PatternFill, named_styles, colors

from babel.dates import format_date
from datetime import date
from copy import copy

from XMLReader import XMLReader
from time import localtime

#Creating 3rd form

class excel_forma:
    # создание объекта
    def __init__(self, path_xml, path_excel):
        self._wb = Workbook()

        self.sec_since_epoch = list(localtime())[:3][::-1]

        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)
        self.font = Font(bold=True, color="000000")
        self.al_center = Alignment(horizontal="center", vertical="center", wrapText = True)
        self.al_left = Alignment(horizontal="left", vertical="center", wrapText = True)

        self._TITLE_FRS_PAGE_ROW = 1 # 1 row
        self._TITLE_FRS_PAGE_COL = 2 # 'B' column

        self._COUNT_COLUMN = 7

        self._TITLE_SCN_PAGE_ROW = self._TITLE_FRS_PAGE_ROW + 2 # 3 row
        self._TITLE_SCN_PAGE_COL = self._TITLE_FRS_PAGE_COL     # 'B' column

        self._TITLE_THR_PAGE_ROW = self._TITLE_SCN_PAGE_ROW + 1 # 4 row
        self._TITLE_THR_PAGE_COL = self._TITLE_FRS_PAGE_COL     # 'B' column

        self._DATE_ROW = self._TITLE_FRS_PAGE_ROW       # 1 row
        self._DATE_COL = self._TITLE_FRS_PAGE_COL + 5   # 'G' column

        self._START_ROW_TABLE = self._TITLE_FRS_PAGE_ROW + 6    # 7 row
        self._START_COL_TABLE = self._TITLE_FRS_PAGE_COL - 1    # 'A' column

        self._TITLE_FRT_PAGE_ROW = self._START_ROW_TABLE        # 7 row
        self._TITLE_FRT_PAGE_COL = self._START_COL_TABLE + 2    # 'C' column


        self.path_excel = path_excel
        self.path_xml = path_xml
    # формирование книги    
    def book_forming(self):
        self._ws_page3 = self._wb.active
        self._ws_page3.title = "Page"        

        self._formatting_page_cells()

        self._filling_page3()

        self.save_file()
    # установка стилей
    def _select_styles_page(self, ws, cell_range, border=None, font=None, alignment=None):
        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = list(ws[cell_range])

        if font:
            first_cell.font = font

        if border:
            for cell in rows[0]:
                cell.border = border
            for cell in rows[-1]:
                cell.border = border

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = border
                r.border = border
    # считывание файла
    def _reading_file(self, path):
        ls_price = XMLReader(path).get_list_of_dicts()
        if ls_price:
            return ls_price 
        else:
            raise Exception("File " + path + " is empty ")
    # форматирование ячеек
    def _formatting_page_cells(self):
        self._ws_page3.column_dimensions[utils.get_column_letter(self._TITLE_FRS_PAGE_COL - 1)].width = 5.43    # 'A' column
        self._ws_page3.column_dimensions[utils.get_column_letter(self._TITLE_FRS_PAGE_COL)].width = 35.57       # 'B' column
        self._ws_page3.column_dimensions[utils.get_column_letter(self._TITLE_FRS_PAGE_COL + 1)].width = 12.29       # 'C' column
        self._ws_page3.column_dimensions[utils.get_column_letter(self._TITLE_FRS_PAGE_COL + 2)].width = 12.29       # 'D' column
        self._ws_page3.column_dimensions[utils.get_column_letter(self._TITLE_FRS_PAGE_COL + 3)].width = 13.14      # 'E' column
        self._ws_page3.column_dimensions[utils.get_column_letter(self._TITLE_FRS_PAGE_COL + 4)].width = 16.86       # 'F' column
        self._ws_page3.column_dimensions[utils.get_column_letter(self._DATE_COL)].width = 24.14       # 'G' column


        self._ws_page3.row_dimensions[self._TITLE_FRS_PAGE_ROW].height = 33    # 1 row
        self._ws_page3.row_dimensions[self._TITLE_SCN_PAGE_ROW].height = 33    # 3 row

        # B1:E1
        self._select_styles_page(self._ws_page3,
                                 utils.get_column_letter(self._TITLE_FRS_PAGE_COL) + str(self._TITLE_FRS_PAGE_ROW) + ':' +
                                 utils.get_column_letter(self._TITLE_FRS_PAGE_COL + 3) + str(self._TITLE_FRS_PAGE_ROW),
                                 font = self.font, alignment = self.al_center)
        # B3:E3
        self._select_styles_page(self._ws_page3,
                                 utils.get_column_letter(self._TITLE_SCN_PAGE_COL) + str(self._TITLE_SCN_PAGE_ROW) + ':' +
                                 utils.get_column_letter(self._TITLE_SCN_PAGE_COL + 3) + str(self._TITLE_SCN_PAGE_ROW),
                                 font = self.font, alignment = self.al_center)
        # B4:E4
        self._select_styles_page(self._ws_page3,
                                 utils.get_column_letter(self._TITLE_SCN_PAGE_COL) + str(self._TITLE_THR_PAGE_ROW) + ':' +
                                 utils.get_column_letter(self._TITLE_SCN_PAGE_COL + 3) + str(self._TITLE_THR_PAGE_ROW),
                                 font = self.font, alignment = self.al_center)

        self._ws_page3[utils.get_column_letter(self._DATE_COL) + str(self._TITLE_FRS_PAGE_ROW)].alignment = self.al_center
        self._ws_page3[utils.get_column_letter(self._DATE_COL) + str(self._TITLE_FRS_PAGE_ROW)].font = self.font

    # заполнение страницы
    def _filling_page3(self):

        product_names = self._reading_file(self.path_xml[3])

        ls_product_names = []
        idx_name_product = 1
               
        for array in product_names:
            ls_product_names.append(array['attrib']['fullname'])
            self._ws_page3.cell(row = 8 + idx_name_product, column = 1, value = idx_name_product).alignment = self.al_center
            self._ws_page3.cell(row = 8 + idx_name_product, column = 2, value = ls_product_names[-1]).alignment = self.al_left
            idx_name_product += 1

        dict_price = self._data_page3_to_dict(ls_product_names)

        dt = self.report_date.split('T')
        date_arr = dt[0].split('-')

        self._ws_page3["B1"] = "Отчет сгенерирован Функциональной подсистемой «Управление сельского хозяйства» акимата города Астаны"
        self._ws_page3["G1"] = "Дата выдачи: " + ".".join(str(el) for el in self.sec_since_epoch)
        self._ws_page3["B3"] = "Ежедневный отчет о мониторинге цен на социально-значимые продовольственные товары"
        self._ws_page3["B4"] = "Отчетный период за " + format_date(date(int(date_arr[0]), int(date_arr[1]), int(date_arr[2])), format = 'long', locale = 'ru_RU')
        
        frs_date = self.first_sunday.split('T')
        rep_date = self.report_date.split('T')
                
        self._ws_page3["A7"] = "№"
        self._ws_page3["B7"] = "Наименование социально-значимых продовольственных товаров"
        self._ws_page3["C7"] = "Цена на " + '.'.join(frs_date[0].split('-')[::-1])
        self._ws_page3["D7"] = "Цена на " + '.'.join(rep_date[0].split('-')[::-1])
        self._ws_page3["E7"] = "Отклонение в % в динамике с " + '.'.join(frs_date[0].split('-')[::-1])
        self._ws_page3["F7"] = "Пороговые значения розничных цен на соц. значимые продовольственные товары на " + self._get_start_quarter(date_arr) + " (утверждены приказом МНЭ)"
        self._ws_page3["G7"] = "Отклонение цены в сравнении с пороговой ценой в %"

        idx = 1
        for col_letter in utils.get_column_interval(1, 7):

            self._ws_page3[col_letter + '8'] = idx
            idx += 1

            for row_number in range(len(ls_product_names) + 2):
                self._ws_page3[col_letter + str(7 + row_number)].border = self.border
                
                self._ws_page3[col_letter + str(7 + row_number)].font = self.font

                if col_letter == 'B' and 9 <= 7 + row_number:
                    self._ws_page3[col_letter + str(7 + row_number)].alignment = self.al_left
                else:
                    self._ws_page3[col_letter + str(7 + row_number)].alignment = self.al_center

        self._filling_table(dict_price, ls_product_names)
    # заполнение таблицы
    def _filling_table(self, dict_price, ls_product_names):
        f = self.font
        for key in dict_price.keys():
            try:
                i_row = ls_product_names.index(key) + 9
            except ValueError:
                pass

            self._ws_page3.cell(row=i_row, column = 3, value = dict_price[key][0]).alignment = self.al_center
            self._ws_page3.cell(row=i_row, column = 4, value = dict_price[key][1]).alignment = self.al_center
            self._ws_page3.cell(row=i_row, column = 6, value = dict_price[key][2]).alignment = self.al_center

            if dict_price[key][0] != '' and dict_price[key][1] != '':
                ls_price_deviation_first_week = round((dict_price[key][1] - dict_price[key][0]) / dict_price[key][0] * 100, 1)
                self._shading_cell(ls_price_deviation_first_week, i_row, 5)
            if dict_price[key][1] != '' and dict_price[key][2] != '':
                ls_price_deviation_from_limit_price = round(dict_price[key][1] * 100 / dict_price[key][2] - 100, 1)
                self._shading_cell(ls_price_deviation_from_limit_price, i_row, 7)
    # раскраска ячеек
    def _shading_cell(self, price, i_row, i_column):
        if price < 0:
            self._ws_page3.cell(row=i_row, column = i_column, value = price).fill = PatternFill(patternType = 'solid', fgColor = colors.GREEN)
        elif price > 0:
            self._ws_page3.cell(row=i_row, column = i_column, value = price).fill = PatternFill(patternType = 'solid', fgColor = colors.RED)
        else:
            self._ws_page3.cell(row=i_row, column = i_column, value = price)
        self._ws_page3.cell(row=i_row, column = i_column, value = price).alignment = self.al_center
    # преобразование данные в (ключ, значение)
    def _data_page3_to_dict(self, ls_product_names):
        
        #список цен на первую неделю
        ls_price_first_week = self._reading_file(self.path_xml[0])

        #список цен на текущую неделю
        ls_price_now_week = self._reading_file(self.path_xml[1])
        
        #пороговые значения цен
        ls_limit_price = self._reading_file(self.path_xml[2])
        
        self.first_sunday = ls_price_first_week[0]['attrib']['Dend']
        self.report_date = ls_price_now_week[0]['attrib']['Dend']

        dict_prod_names = {}
        
        idx = 0
        for product_name in ls_product_names:
            
            for ls_first_week in ls_price_first_week:
                if ls_first_week['attrib']['item2'] == product_name:
                    if dict_prod_names.get(product_name):
                        dict_prod_names[product_name].append(int(float(ls_first_week['attrib']['Value1'])))
                    else:
                        dict_prod_names[product_name] = [int(float(ls_first_week['attrib']['Value1']))]
                    break

            for ls_now_week in ls_price_now_week:
                if ls_now_week['attrib']['item2'] == product_name:
                    if dict_prod_names.get(product_name):
                        dict_prod_names[product_name].append(int(float(ls_now_week['attrib']['Value1'])))
                    else:
                        dict_prod_names[product_name] = ['', int(float(ls_now_week['attrib']['Value1']))]
                    break

            if dict_prod_names.get(product_name) and 1 == len(dict_prod_names[product_name]):
                dict_prod_names[product_name].append('')

            for limit_price in ls_limit_price:
                if limit_price['attrib']['item2'] == product_name:
                    if dict_prod_names.get(product_name):
                        dict_prod_names[product_name].append(int(float(limit_price['attrib']['Value1'])))
                    else:
                        dict_prod_names[product_name] = ['', '', int(float(limit_price[65]))]
                    break  

            if dict_prod_names.get(product_name) and 2 == len(dict_prod_names[product_name]):
                dict_prod_names[product_name].append('')

            if dict_prod_names.get(product_name) is None:
                dict_prod_names[product_name] = ['', '', '']
        return dict_prod_names
    # квартал по дате
    def _get_start_quarter(self, dt_end):
        if int(dt_end[1]) <= 3:
            return "1 кв. " + dt_end[0] + ' г.'
        elif int(dt_end[1]) <= 6:
            return "2 кв. " + dt_end[0] + ' г.'
        if int(dt_end[1]) <= 9:
            return "3 кв. " + dt_end[0] + ' г.'
        if int(dt_end[1]) <= 12:
            return "4 кв. " + dt_end[0] + ' г.'
    # сохранение в файл
    def save_file(self): 
        self._wb.save(self.path_excel)
        print("true")