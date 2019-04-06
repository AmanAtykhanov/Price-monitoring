from openpyxl import Workbook, utils, worksheet
from openpyxl.styles import Fill, Font, Border, Side, Alignment, PatternFill, named_styles, colors

import time
from babel.dates import format_date
from datetime import date
from copy import copy

import XMLReader as xr

#Creating 3rd form

class excel_forma:

    def __init__(self, path_xml, path_excel):
        self._wb = Workbook()

        self.sec_since_epoch = list(time.localtime())[:3][::-1]

        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)
        self.font = Font(bold=True, color="000000")
        self.al = Alignment(horizontal="center", vertical="center", wrapText = True)

        self.path_excel = path_excel
        self.path_xml = path_xml
        
    def page3_forming(self):
        self._ws_page3 = self._wb.active
        self._ws_page3.title = "Page"        

        self._merge_page3_cells()

        self._filling_page3()

        self.save_file()

    def _select_styles_page(self, ws, cell_range, border=None, font=None, alignment=None):
        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = list(ws[cell_range])

        if font:
            first_cell.font = font

        if border:
            for cell in rows[0]:
                cell.border = border
            for cell in rows[-1]:
                cell.border = border

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = border
                r.border = border

    def _merge_page3_cells(self):
        self._ws_page3.column_dimensions['A'].width = 5.43
        self._ws_page3.column_dimensions['B'].width = 35.57
        self._ws_page3.column_dimensions['C'].width = 12.29
        self._ws_page3.column_dimensions['D'].width = 12.29
        self._ws_page3.column_dimensions['E'].width = 13.14
        self._ws_page3.column_dimensions['F'].width = 16.86
        self._ws_page3.column_dimensions['G'].width = 24.14


        self._ws_page3.row_dimensions[1].height = 33
        self._ws_page3.row_dimensions[3].height = 37.50

        self._select_styles_page(self._ws_page3, "B1:E1", font = self.font, alignment = self.al)
        self._select_styles_page(self._ws_page3, "B3:E3", font = self.font, alignment = self.al)
        self._select_styles_page(self._ws_page3, "B4:E4", font = self.font, alignment = self.al)
        self._ws_page3['G1'].font = self.font
        self._ws_page3['G1'].alignment = self.al

    def _filling_page3(self):

        xml_obj = xr.XMLReader(self.path_xml[3])
        product_names = xml_obj.get_list_of_dicts()

        ls_product_names = []
        idx_name_product = 1

        align = Alignment(horizontal="left", vertical="center", wrapText = True)
        
        for array in product_names:
            ls_product_names.append(array['attrib']['fullname'])
            self._ws_page3.cell(row = 8 + idx_name_product, column = 1, value = idx_name_product).alignment = self.al
            self._ws_page3.cell(row = 8 + idx_name_product, column = 2, value = ls_product_names[-1]).alignment = align
            idx_name_product += 1

        dict_price = self._data_page3_to_dict(ls_product_names)

        dt = self.report_date.split('T')
        date_arr = dt[0].split('-')

        self._ws_page3["B1"] = "Отчет сгенерирован Функциональной подсистемой «Управление сельского хозяйства» акимата города Астаны"
        self._ws_page3["G1"] = "Дата выдачи: " + ".".join(str(el) for el in self.sec_since_epoch)
        self._ws_page3["B3"] = "Ежедневный отчет о мониторинге цен на социально-значимые продовольственные товары"
        self._ws_page3["B4"] = "Отчетный период за " + format_date(date(int(date_arr[0]), int(date_arr[1]), int(date_arr[2])), format = 'long', locale = 'ru_RU')
        
        frs_date = self.first_sunday.split('T')
        rep_date = self.report_date.split('T')
                
        self._ws_page3["A7"] = "№"
        self._ws_page3["B7"] = "Наименование социально-значимых продовольственных товаров"
        self._ws_page3["C7"] = "Цена на " + '.'.join(frs_date[0].split('-')[::-1])
        self._ws_page3["D7"] = "Цена на " + '.'.join(rep_date[0].split('-')[::-1])
        self._ws_page3["E7"] = "Отклонение в % в динамике с " + '.'.join(frs_date[0].split('-')[::-1])
        self._ws_page3["F7"] = "Пороговые значения розничных цен на соц. значимые продовольственные товары на " + self._get_start_quarter(date_arr) + " (утверждены приказом МНЭ)"
        self._ws_page3["G7"] = "Отклонение цены в сравнении с пороговой ценой в %"

        idx = 1
        for col_letter in utils.get_column_interval(1, 7):

            self._ws_page3[col_letter + '8'] = idx
            idx += 1

            for row_number in range(len(ls_product_names) + 2):
                self._ws_page3[col_letter + str(7 + row_number)].border = self.border
                
                self._ws_page3[col_letter + str(7 + row_number)].font = self.font

                if col_letter == 'B' and 9 <= 7 + row_number:
                    self._ws_page3[col_letter + str(7 + row_number)].alignment = align
                else:
                    self._ws_page3[col_letter + str(7 + row_number)].alignment = self.al

        self._filling_value_page3(dict_price, ls_product_names)

    def _filling_value_page3(self, dict_price, ls_product_names):
        f = self.font
        for key in dict_price.keys():
            try:
                i_row = ls_product_names.index(key) + 9
            except ValueError:
                pass

            self._ws_page3.cell(row=i_row, column = 3, value = dict_price[key][0]).alignment = self.al
            self._ws_page3.cell(row=i_row, column = 4, value = dict_price[key][1]).alignment = self.al
            self._ws_page3.cell(row=i_row, column = 6, value = dict_price[key][2]).alignment = self.al

            if dict_price[key][0] != '' and dict_price[key][1] != '':
                ls_price_deviation_first_week = round((dict_price[key][1] - dict_price[key][0]) / dict_price[key][0] * 100, 1)
                self._shading_cell(ls_price_deviation_first_week, i_row, 5)
            if dict_price[key][1] != '' and dict_price[key][2] != '':
                ls_price_deviation_from_limit_price = round(dict_price[key][1] * 100 / dict_price[key][2] - 100, 1)
                self._shading_cell(ls_price_deviation_from_limit_price, i_row, 7)

    def _shading_cell(self, price, i_row, i_column):
        if price < 0:
            self._ws_page3.cell(row=i_row, column = i_column, value = price).fill = PatternFill(patternType = 'solid', fgColor = colors.GREEN)
        elif price > 0:
            self._ws_page3.cell(row=i_row, column = i_column, value = price).fill = PatternFill(patternType = 'solid', fgColor = colors.RED)
        else:
            self._ws_page3.cell(row=i_row, column = i_column, value = price)
        self._ws_page3.cell(row=i_row, column = i_column, value = price).alignment = self.al

    def _data_page3_to_dict(self, ls_product_names):
        
        #список цен на первую неделю
        xml_obj = xr.XMLReader(self.path_xml[0])
        ls_price_first_week = xml_obj.get_list_of_dicts()

        #список цен на текущую неделю
        xml_obj = xr.XMLReader(self.path_xml[1])
        ls_price_now_week = xml_obj.get_list_of_dicts()
        
        #отклонение цен на текущую неделю
        #xml_obj = xr.XMLReader(self.path_xml[2])
        #ls_price_deviation_first_week = xml_obj.get_list_of_dicts()
        
        #пороговые значения цен
        xml_obj = xr.XMLReader(self.path_xml[2])
        ls_limit_price = xml_obj.get_list_of_dicts()     

        #отклонение цены в сравнении с пороговой ценой
        #xml_obj = xr.XMLReader(self.path_xml[4])
        #ls_price_deviation_from_limit_price = xml_obj.get_list_of_dicts()
        
        self.first_sunday = ls_price_first_week[0]['attrib']['Dend']
        self.report_date = ls_price_now_week[0]['attrib']['Dend']

        dict_prod_names = {}
        
        idx = 0
        for product_name in ls_product_names:
            
            for ls_first_week in ls_price_first_week:
                if ls_first_week['attrib']['item2'] == product_name:
                    if dict_prod_names.get(product_name):
                        dict_prod_names[product_name].append(int(float(ls_first_week['attrib']['Value1'])))
                    else:
                        dict_prod_names[product_name] = [int(float(ls_first_week['attrib']['Value1']))]
                    break

            for ls_now_week in ls_price_now_week:
                if ls_now_week['attrib']['item2'] == product_name:
                    if dict_prod_names.get(product_name):
                        dict_prod_names[product_name].append(int(float(ls_now_week['attrib']['Value1'])))
                    else:
                        dict_prod_names[product_name] = ['', int(float(ls_now_week['attrib']['Value1']))]
                    break

            if dict_prod_names.get(product_name) and 1 == len(dict_prod_names[product_name]):
                dict_prod_names[product_name].append('')

            #for ls_first_week in ls_price_deviation_first_week:
            #    if ls_first_week['attrib']['item2'] == product_name:
            #        if dict_prod_names.get(product_name):
            #            dict_prod_names[product_name].append(float(ls_first_week['attrib']['Value1']))
            #        else:
            #            dict_prod_names[product_name] = ['', '', float(ls_first_week['attrib']['Value1'])]
            #        break

            #if dict_prod_names.get(product_name) and 2 == len(dict_prod_names[product_name]):
            #    dict_prod_names[product_name].append('')

            for limit_price in ls_limit_price:
                if limit_price['attrib']['item2'] == product_name:
                    if dict_prod_names.get(product_name):
                        dict_prod_names[product_name].append(int(float(limit_price['attrib']['Value1'])))
                    else:
                        dict_prod_names[product_name] = ['', '', int(float(limit_price[65]))]
                    break  

            if dict_prod_names.get(product_name) and 2 == len(dict_prod_names[product_name]):
                dict_prod_names[product_name].append('')

            #for deviation_from_limit_price in ls_price_deviation_from_limit_price:
            #    if deviation_from_limit_price['attrib']['item2'] == product_name:
            #        if dict_prod_names.get(product_name):
            #            dict_prod_names[product_name].append(float(deviation_from_limit_price['attrib']['Value1']))
            #        else:
            #            dict_prod_names[product_name] = ['', '', '', '', float(deviation_from_limit_price['attrib']['Value1'])]
            #        break
            if dict_prod_names.get(product_name) is None:
                dict_prod_names[product_name] = ['', '', '']
        return dict_prod_names
        

    def _get_start_quarter(self, dt_end):
        if int(dt_end[1]) <= 3:
            return "1 кв. " + dt_end[0] + ' г.'
        elif int(dt_end[1]) <= 6:
            return "2 кв. " + dt_end[0] + ' г.'
        if int(dt_end[1]) <= 9:
            return "3 кв. " + dt_end[0] + ' г.'
        if int(dt_end[1]) <= 12:
            return "4 кв. " + dt_end[0] + ' г.'

    def save_file(self): 
        self._wb.save(self.path_excel)
        print("true")