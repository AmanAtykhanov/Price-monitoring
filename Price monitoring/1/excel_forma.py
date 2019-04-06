from openpyxl import Workbook, utils
from openpyxl.styles import Font, Border, Side, Alignment, named_styles

from XMLReader import XMLReader

import time
from babel.dates import format_date
from datetime import date




class excel_forma:
    
    def __init__(self, path_xml, path_excel):
        self._wb = Workbook()
        self.sec_since_epoch = list(time.localtime())[:3][::-1]

        self.thin = Side(border_style="thin", color="000000")
        self.border = Border(top=self.thin, left=self.thin, right=self.thin, bottom=self.thin)
        self.font = Font(bold=True, color="000000")
        self.al = Alignment(horizontal="center", vertical="center", wrapText = True)

        self.path_xml = path_xml
        self.path_excel = path_excel

    def page1_forming(self):
        self._ws_page1 = self._wb.active
        self._ws_page1.title = "Page"
        dict_data = self._data2dict()
        self._filling_page1(dict_data['Market'])
        self._merge_page1_cells()

        self.save_file()

    def _reading_file(self, path):
        ls_avr_sum = XMLReader(path).get_list_of_dicts()
        if ls_avr_sum:
            return ls_avr_sum 
        else:
            raise Exception("File " + path + " is empty ")

    def _reading_avr_sum(self):
        dict_market_avr_sum = {'Market': {}}

        ls_avr_sum = self._reading_file(self.path_xml[0])

        self.report_date = ls_avr_sum[0]['attrib']['Dend']
        for avr_sum in ls_avr_sum:
            if dict_market_avr_sum['Market'].get(avr_sum['attrib']['item3']):
                if dict_market_avr_sum['Market'][avr_sum['attrib']['item3']].get('product'):
                    dict_market_avr_sum['Market'][avr_sum['attrib']['item3']]['product'].append(avr_sum['attrib']['item2'])
                else:
                    dict_market_avr_sum['Market'][avr_sum['attrib']['item3']]['product'] = [avr_sum['attrib']['item2']]
            else:
                dict_market_avr_sum['Market'][avr_sum['attrib']['item3']] = {'product': [avr_sum['attrib']['item2']]}


            if dict_market_avr_sum['Market'].get(avr_sum['attrib']['item3']):
                if dict_market_avr_sum['Market'][avr_sum['attrib']['item3']].get('avr_sum'):
                    dict_market_avr_sum['Market'][avr_sum['attrib']['item3']]['avr_sum'].append(int(float(avr_sum['attrib']['Value1'])))
                else:
                    dict_market_avr_sum['Market'][avr_sum['attrib']['item3']]['avr_sum'] = [int(float(avr_sum['attrib']['Value1']))]
            else:
                dict_market_avr_sum['Market'][avr_sum['attrib']['item3']] = {'avr_sum': [int(float(avr_sum['attrib']['Value1']))]}
        
        return dict_market_avr_sum
    
    def _reading_minmax_sum(self):
        dict_market_minmax_sum = {'Market': {}}

        ls_minmax_sum = self._reading_file(self.path_xml[1])

        for minmax_sum in ls_minmax_sum:
            #min costs
            if 'min' == minmax_sum['attrib']['item1']:
                if dict_market_minmax_sum['Market'].get(minmax_sum['attrib']['item4']):
                    if dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']].get('min'):
                        dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']]['min'].append(int(float(minmax_sum['attrib']['Value1'])))
                    else:
                        dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']]['min'] = [int(float(minmax_sum['attrib']['Value1']))]
                else:
                    dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']] = {'min': [int(float(minmax_sum['attrib']['Value1']))]}
            #max costs
            elif 'max' == minmax_sum['attrib']['item1']:
                if dict_market_minmax_sum['Market'].get(minmax_sum['attrib']['item4']):
                    if dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']].get('max'):
                        dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']]['max'].append(int(float(minmax_sum['attrib']['Value1'])))
                    else:
                        dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']]['max'] = [int(float(minmax_sum['attrib']['Value1']))]
                else:
                    dict_market_minmax_sum['Market'][minmax_sum['attrib']['item4']] = {'max': [int(float(minmax_sum['attrib']['Value1']))]}

        return dict_market_minmax_sum

    def _select_styles_page(self, ws, cell_range, border=None, font=None, alignment=None):
        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = list(ws[cell_range])

        if font:
            first_cell.font = font

        if border:
            for cell in rows[0]:
                cell.border = border
            for cell in rows[-1]:
                cell.border = border

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = border
                r.border = border

    def _merge_page1_cells(self):
        self._ws_page1.column_dimensions['A'].width = 5.43
        self._ws_page1.column_dimensions['B'].width = 35.57

        self._ws_page1.row_dimensions[1].height = 33
        self._ws_page1.row_dimensions[7].height = 32.25
        self._ws_page1.row_dimensions[3].height = 39.75
        self._select_styles_page(self._ws_page1, "B1:E1", font = self.font, alignment = self.al)
        self._select_styles_page(self._ws_page1, "B3:E3", font = self.font, alignment = self.al)
        self._select_styles_page(self._ws_page1, "B4:E4", font = self.font, alignment = self.al)
        self._select_styles_page(self._ws_page1, "A7:A8", font = self.font, alignment = self.al)
        self._select_styles_page(self._ws_page1, "G1:I1", font = self.font, alignment = self.al)

    def _filling_page1(self, dict_markets):
        dt = self.report_date.split('T')
        dt = dt[0].split('-')
        self._ws_page1["B4"] = "отчетный период за " + format_date(date(int(dt[0]), int(dt[1]), int(dt[2])), format = 'long', locale = 'ru_RU')
        self._ws_page1["B1"] = "Отчет сгенерирован Функциональной подсистемой «Управление сельского хозяйства» акимата города Астаны"
        self._ws_page1["B3"] = "Отчет о мониторинге цен на социально-значимые продовольственные товары в разрезе торговых объектов"
        self._ws_page1["G1"] = "Дата выдачи: " + ".".join(str(el) for el in self.sec_since_epoch)
        self._ws_page1["A7"] = "№"

        product_names = self._reading_file(self.path_xml[2])

        self._ws_page1.cell(row = 9, column = 1, value = 1)
        self._ws_page1.cell(row = 9, column = 2, value = 2)
        ls_product_names = []
        idx_name_product = 1


        align = Alignment(horizontal="left", vertical="center", wrapText = True)

        for array in product_names:
            ls_product_names.append(array['attrib']['fullname'])
            self._ws_page1.cell(row = 9 + idx_name_product, column = 1, value = idx_name_product)
            self._ws_page1.cell(row = 9 + idx_name_product, column = 2, value = ls_product_names[-1])

            idx_name_product += 1

        
        for i_col in range(len(dict_markets)):
            self._ws_page1.merge_cells(start_row = 7, start_column = i_col * 3 + 3, end_row = 7, end_column = i_col * 3 + 5)
        i_col = 0
        for name_market in dict_markets:
            
            #name market
            self._ws_page1.cell(row = 7, column = i_col * 3 + 3, value = name_market).border = self.border
            #name min
            self._ws_page1.cell(row = 8, column = i_col * 3 + 3, value = 'min')
            self._ws_page1.cell(row = 9, column = i_col * 3 + 3, value = i_col * 3 + 3)
            #name max
            self._ws_page1.cell(row = 8, column = i_col * 3 + 4, value = 'max')
            self._ws_page1.cell(row = 9, column = i_col * 3 + 4, value = i_col * 3 + 4)
            #name avr_sum
            self._ws_page1.cell(row = 8, column = i_col * 3 + 5, value = 'Ср.цена')
            self._ws_page1.cell(row = 9, column = i_col * 3 + 5, value = i_col * 3 + 5)
           
            self._filling_value_page1(dict_markets[name_market], i_col * 3 + 3, ls_product_names)
            i_col += 1

        for col_letter in utils.get_column_interval(1, len(dict_markets) * 3 + 2):
            for row_number in range(len(ls_product_names) + 3):
                self._ws_page1[col_letter + str(7 + row_number)].font = self.font
                self._ws_page1[col_letter + str(7 + row_number)].border = self.border
                if col_letter == 'B' and 10 <= 7 + row_number:
                    self._ws_page1[col_letter + str(7 + row_number)].alignment = align
                else:
                    self._ws_page1[col_letter + str(7 + row_number)].alignment = self.al

    def _filling_value_page1(self, market, i_col, ls_product_names):
        length = len(market['min'])
            
        for count in range(length):
            try:
                i_row = ls_product_names.index(market['product'][count]) + 10
            except ValueError:
                continue
            local_index_col = i_col

            self._ws_page1.cell(row = i_row, column = local_index_col, value = market['min'][count])
            local_index_col += 1

            self._ws_page1.cell(row = i_row, column = local_index_col, value = market['max'][count])
            local_index_col += 1

            self._ws_page1.cell(row = i_row, column = local_index_col, value = market['avr_sum'][count])
            count += 1

    def _data2dict(self):
        dict_avr_sum = self._reading_avr_sum()
        dict_minmax_sum = self._reading_minmax_sum()
        dict_cost = dict_avr_sum.copy()
        
        for key in dict_avr_sum['Market'].keys():
            dict_cost['Market'][key].update(dict_minmax_sum['Market'][key])
        return dict_cost

    def save_file(self): 
        self._wb.save(self.path_excel)
        print("true")